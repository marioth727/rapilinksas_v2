# -*- coding: utf-8 -*-
import sys
import os
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

# Find the file dynamically
base = r'C:\Users\Usuario\OneDrive\Documentos\desarrollo antgra\Pagina Rapilink\recursos'
files = os.listdir(base)
print("Files in recursos:", files)

xlsx = [f for f in files if f.endswith('.xlsx')][0]
path = os.path.join(base, xlsx)
print("Opening:", path)

wb = openpyxl.load_workbook(path)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f'\n=== HOJA: {sheet} ===')
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            # Print each cell safely
            cells = [str(c)[:200] if c else '' for c in row]
            print(' | '.join(cells))
